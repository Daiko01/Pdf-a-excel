# -*- coding: utf-8 -*-
from __future__ import annotations
import logging
from typing import List, Dict, Any
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from schema import ROW_SCHEMA, TYPE_CASTERS

LOGGER = logging.getLogger("excel_io")
SHEET_NAME = "Datos"

def ensure_schema_columns(df: pd.DataFrame) -> pd.DataFrame:
    for col in ROW_SCHEMA:
        if col not in df.columns:
            df[col] = None
    return df[ROW_SCHEMA]

def cast_types(df: pd.DataFrame) -> pd.DataFrame:
    for col, caster in TYPE_CASTERS.items():
        if col in df.columns:
            df[col] = df[col].map(caster)
    return df

def read_excel_all_sheets(path: str | Path) -> Dict[str, pd.DataFrame]:
    path = Path(path)
    if not path.exists():
        return {}
    wb = load_workbook(path)
    data: Dict[str, pd.DataFrame] = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.values)
        if not rows:
            data[name] = pd.DataFrame()
            continue
        header = list(rows[0])
        body = list(rows[1:])
        df = pd.DataFrame(body, columns=header)
        data[name] = df
    return data

def write_preserving_other_sheets(path: str | Path, df_datos: pd.DataFrame) -> None:
    path = Path(path)
    existing = read_excel_all_sheets(path) if path.exists() else {}
    existing[SHEET_NAME] = ensure_schema_columns(df_datos.copy())
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, df in existing.items():
            if df is None:
                df = pd.DataFrame()
            df = ensure_schema_columns(df) if name == SHEET_NAME else df
            df.to_excel(writer, sheet_name=name, index=False)

def append_and_dedup(base_path: str | Path, new_rows: List[Dict[str, Any]], out_path: str | Path | None = None) -> str:
    base_path = Path(base_path)
    if out_path is None:
        out_path = base_path
    out_path = Path(out_path)

    all_sheets = read_excel_all_sheets(base_path) if base_path.exists() else {}
    df_old = all_sheets.get(SHEET_NAME, pd.DataFrame(columns=ROW_SCHEMA))
    df_old = ensure_schema_columns(df_old)

    df_new = pd.DataFrame(new_rows, columns=ROW_SCHEMA)
    df_new = ensure_schema_columns(df_new)

    df_concat = pd.concat([df_old, df_new], ignore_index=True)
    df_concat = df_concat.drop_duplicates(subset=["Folio","Fecha","Máquina"], keep="first")
    df_concat = cast_types(df_concat)

    write_preserving_other_sheets(out_path, df_concat)
    return str(out_path)

def create_new_excel(out_path: str | Path, rows: List[Dict[str, Any]]) -> str:
    out_path = Path(out_path)
    df = pd.DataFrame(rows, columns=ROW_SCHEMA)
    df = ensure_schema_columns(df)
    df = cast_types(df)
    write_preserving_other_sheets(out_path, df)
    return str(out_path)
